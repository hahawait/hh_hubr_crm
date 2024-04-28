from io import BytesIO
from openpyxl import Workbook

from apps.hubr import models


def create_company_contacts_members_excel(companies_contacts_members):
    # Создание нового файла XLSX в памяти
    wb = Workbook()
    ws = wb.active

    # Заголовки
    headers = [
        "Название компании",
        "Имя сотрудника",
        "Телефон",
        "Email",
        "Telegram",
        "Вконтакте",
        "Facebook",
        "LinkedIn",
        "Twitter"
    ]
    ws.append(headers)

    # Заполнение данных
    for company_contact_members in companies_contacts_members:
        for contact_member in company_contact_members.contact_members:
            # Проверка наличия "mail" в значении для поля email
            row_data = [
                company_contact_members.company_name,
                contact_member.contact_name,
                contact_member.contacts.get("Телефон", ""),
                contact_member.contacts.get("Почта", ""),
                contact_member.contacts.get("Telegram", ""),
                contact_member.contacts.get("Вконтакте", ""),
                contact_member.contacts.get("Facebook", ""),
                contact_member.contacts.get("LinkedIn", ""),
                contact_member.contacts.get("Twitter", "")
            ]
            ws.append(row_data)

    # Сохранение данных в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_company_contacts_excel(companies_contacts: list[models.CompanyModel]) -> BytesIO:
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Настройка заголовков колонок
    headers = [
        "Название компании",
        "Телефон",
        "Email",
        "Telegram",
        "Вконтакте",
        "Facebook",
        "LinkedIn",
        "Twitter"
    ]
    ws.append(headers)

    # Заполняем данные в книгу
    for company_contact in companies_contacts:
        # Создаем список для хранения контактных данных в правильном порядке
        contact_data = [company_contact.company_name]

        # Перебираем типы контактов и добавляем их в список
        for contact_type in headers[1:]:
            contact_data.append(company_contact.contacts.get(contact_type, ""))

        ws.append(contact_data)

    # Сохраняем книгу в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Перемещаем указатель в начало буфера

    return buffer


def create_vacancy_excel(vacancies: list[models.VacancyModel]) -> BytesIO:
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Настройка заголовков колонок
    headers = [
        "Название компании",
        "Название вакансии",
        "Зарплата",
        "Дата"
    ]

    ws.append(headers)

    # Заполняем данные в книгу
    for vacancy in vacancies:
        ws.append([
            vacancy.company_name,
            vacancy.vacancy_name,
            vacancy.salary,
            vacancy.date.strftime("%Y-%m-%d %H:%M:%S") if vacancy.date else None
        ])

    # Сохраняем книгу в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Перемещаем указатель в начало буфера

    return buffer